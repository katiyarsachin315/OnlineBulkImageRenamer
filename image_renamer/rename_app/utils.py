from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
import shutil
import zipfile
import os
from django.http import HttpResponse
import re
import pandas as pd


def validate_image_extensions(uploaded_files):
    valid_extensions = ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'tiff', 'webp']
    error_files = []

    for uploaded_file in uploaded_files:
        file_extension = os.path.splitext(uploaded_file.name)[1][1:].lower()

        if file_extension not in valid_extensions:
            print(f"Unsupported file extension '{file_extension}' in file '{uploaded_file.name}'.")
            error_files.append(uploaded_file.name)

    if error_files:
        return {
            "status": "error",
            "message": f"Unsupported file extension. Errors occurred for files: {', '.join(error_files)}"
        }

    return {"status": "success", "message": "All files are valid."}

def create_folder(path, folder_name):
    full_path = os.path.join(path, folder_name)
    if not os.path.exists(full_path):
        os.makedirs(full_path)
        print(f"Folder '{folder_name}' created at '{path}'.")
    else:
        print(f"Folder '{folder_name}' already exists at '{path}'.")


def get_files_in_folder(folder_path, allow_extensions=None):
    if not os.path.exists(folder_path):
        return {}

    if allow_extensions is None:
        allow_extensions = []

    files_dict = {}
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path):
            if not allow_extensions or any(filename.lower().endswith(ext.lower()) for ext in allow_extensions):
                files_dict[filename] = filename

    return files_dict


import os
import shutil

def rename_file_names(src_path, rename_dict, allowed_extensions=None):
    try:
        # Ensure the provided path exists and is a directory
        if not os.path.exists(src_path):
            raise ValueError(f"The path {src_path} does not exist.")
        if not os.path.isdir(src_path):
            raise ValueError(f"The path {src_path} is not a directory.")

        # If no allowed_extensions are provided, default to an empty list
        if allowed_extensions is None:
            allowed_extensions = ['.jpg', '.jpeg', '.png']

        # Create the 'renamed_files' folder inside the source path if it doesn't exist
        renamed_files_path = os.path.join(src_path, 'renamed_files')
        if not os.path.exists(renamed_files_path):
            os.makedirs(renamed_files_path)
            print(f"Created directory: {renamed_files_path}")

        # Clean the 'renamed_files' folder by deleting existing files (if any)
        existing_files = [f for f in os.listdir(renamed_files_path) if os.path.isfile(os.path.join(renamed_files_path, f))]
        for file in existing_files:
            file_path = os.path.join(renamed_files_path, file)
            os.remove(file_path)
            print(f"Deleted existing file: {file}")

        # List all files in the source directory
        files = [f for f in os.listdir(src_path) if os.path.isfile(os.path.join(src_path, f))]

        # Rename files based on the provided dictionary
        success_files = []
        skipped_files = []

        for file in files:
            # Check if the file has an allowed extension
            file_extension = os.path.splitext(file)[1].lower()
            if file_extension in allowed_extensions:
                if file in rename_dict:
                    # Rename and copy the file
                    new_file_name = rename_dict[file]
                    old_file_path = os.path.join(src_path, file)
                    new_file_path = os.path.join(renamed_files_path, new_file_name)
                    shutil.copy(old_file_path, new_file_path)
                    print(f"Renamed and copied: {file} to {new_file_name} in {renamed_files_path}")
                    success_files.append(file)
                else:
                    # Skip files that are not in the rename_dict
                    print(f"Skipping file as it does not have an entry in rename_dict: {file}")
                    skipped_files.append(file)
            else:
                # Skipping unsupported files
                print(f"Skipping file with unsupported extension: {file}")
                skipped_files.append(file)

        # Return success or message about skipped files
        if len(success_files) > 0:
            return {
                "status": "success",
                "message": f"All the images successfully renamed"
            }
        else:
            return {
                "status": "success",
                "message": "No images were renamed, as no matching files were found in rename_dict."
            }

    except Exception as e:
        print(f"An error occurred: {e}")
        return {
            "status": "error",
            "message": f"An unexpected error occurred: {str(e)}"
        }


        
def zip_folder(folder_path, zip_name):
    zip_path = os.path.join(folder_path, zip_name)
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                # Skip the zip file itself
                if file_path == zip_path:
                    continue
                zipf.write(file_path, os.path.relpath(file_path, folder_path))
    
    return zip_path

def clean_images_name(uploaded_files_name):
        uploaded_files_name_clean={}
        text_to_replace = '_'
        for key, value in uploaded_files_name.items():
            cleaned_name = re.sub(r'[^a-zA-Z0-9.]', text_to_replace, re.sub(r'[\[\]\(\)\{\}]', '', value))
            cleaned_name = re.sub(r'_+', '_', cleaned_name)
            uploaded_files_name_clean[key] = cleaned_name
        return uploaded_files_name_clean


def WriteDataFile(data, path, sheet_name):
    # Convert the data dictionary to a DataFrame
    df = pd.DataFrame(data)
    
    # Write the DataFrame to Excel
    df.to_excel(path, sheet_name=sheet_name, startrow=0, startcol=0, index=False)
    
    # Load the workbook to adjust column widths
    wb = load_workbook(path)
    
    # Loop through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Adjust the column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width
    
    # Save the workbook after adjustments
    wb.save(path)
    
def create_excel_file(data_dict, output_path, base_url):
    files_dict = {'Original_Image_Name': [], 'Cleaned_Image_Name': []}
    
    for key, value in data_dict.items():
        files_dict['Original_Image_Name'].append(key)
        files_dict['Cleaned_Image_Name'].append(value)
    
    path = os.path.join(output_path, 'Image_Name.xlsx')
    
    # Assuming WriteDataFile is your custom Excel writer class
    WriteDataFile(files_dict, path, 'Images-Name')
    
    # Extract the relative path to 'media' folder from output_path
    rel_path = os.path.relpath(path, 'media')
    
    excel_file_path = f"{base_url}media/{rel_path.replace(os.sep, '/')}"
    
    aTag = f'<a href="{excel_file_path}" target="_blank" class="btn download-excelBtn" download>Download File</a>'
    return excel_file_path


def excel_to_dict(file_path):
    # Read the Excel file without treating any row as the header
    data_df = pd.read_excel(file_path, sheet_name=0, header=None)
    
    # Initialize an empty dictionary to store the result
    result_dict = {}
    
    # Iterate over the rows, skipping the first one (header)
    for i in range(1, len(data_df)):
        key = data_df.iloc[i, 0]  # First column value (key)
        value = data_df.iloc[i, 1]  # Second column value (value)
        result_dict[key] = value
    
    # Return the result dictionary
    return result_dict


def rename_files_with_given_list(images_name_array, uploaded_files_name_clean):
    matchedImageName = []

    # Iterate over the image names
    for img_name in images_name_array:
        img_name_val = img_name.split('.')[0]  # Extract the name without the extension
        
        # Iterate over the keys in the uploaded_files_name_clean dictionary
        for key, key_name in uploaded_files_name_clean.items():
            key_name = str(key_name).split('.')[0]  # Extract the key name without the extension
            
            if img_name_val in key_name:  # If the image name is found in the key name
                matchedImageName.append(img_name_val)
                increment_variable = images_name_array.index(img_name) + 1
                file_extension = 'png'
                
                print("matchedImageName:", matchedImageName)
                print("increment_variable:", increment_variable)
                print("file_extension:", file_extension)
                print("key_name matched:", key_name)
                
                imgCount = matchedImageName.count(img_name_val)
                print("imgCount:", imgCount)
                
                # Update the dictionary based on the count of occurrences
                if imgCount > 1:
                    uploaded_files_name_clean[key] = f"{increment_variable}_{imgCount-1}_{img_name_val}.{file_extension}"
                else:
                    uploaded_files_name_clean[key] = f"{increment_variable}_{img_name_val}.{file_extension}"

    return uploaded_files_name_clean

# def delete_folder_after_download(folder_path):
#     try:
#         if os.path.exists(folder_path):
#             shutil.rmtree(folder_path)  # Delete the folder
#             print(f"Deleted folder: {folder_path}")
#         else:
#             print(f"Folder {folder_path} does not exist")
#     except Exception as e:
#         print(f"Error deleting folder: {e}")